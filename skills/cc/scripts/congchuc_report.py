#!/usr/bin/env python3
"""Export Excel (.xlsx) và báo cáo thống kê công văn đến.

Usage:
  congchuc_report.py --excel              -> sinh file .xlsx
  congchuc_report.py --weekly             -> báo cáo tuần (stdout -> Zalo)
  congchuc_report.py --monthly            -> báo cáo tháng
  congchuc_report.py --check-overdue      -> tự động đánh dấu quá hạn
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import json, os, datetime

_hermes_home = os.environ.get("HERMES_HOME", os.path.expanduser("~/.hermes"))
STATE_FILE = os.path.join(_hermes_home, "cron", "cong-van-den", "vbden_state.json")
EXPORT_DIR = os.path.join(_hermes_home, "cron", "cong-van-den", "exports")

URGENT_KEYWORDS = ['Cực Khẩn', 'Hỏa tốc hẹn giờ', 'Hỏa tốc', 'Thượng khẩn', 'Khẩn', 'Gấp', 'Tốc ký']
OVERDUE_DAYS = int(os.environ.get("CONGVAN_OVERDUE_DAYS", "7"))


def load_state():
    if os.path.exists(STATE_FILE):
        with open(STATE_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {"seen_ids": [], "documents": {}, "last_check": None, "last_count": 0}


def save_state(state):
    os.makedirs(os.path.dirname(STATE_FILE), exist_ok=True)
    with open(STATE_FILE, 'w', encoding='utf-8') as f:
        json.dump(state, f, indent=2, ensure_ascii=False)


def get_urgency(doc):
    parsed = doc.get('do_khan', '').strip()
    if parsed and parsed != 'Thường':
        return parsed
    combined = f"{doc.get('trich_yeu', '')} {doc.get('so_ky_hieu', '')} {doc.get('so_den', '')}"
    combined = combined.replace('(', '').replace(')', '')
    for kw in URGENT_KEYWORDS:
        if kw.lower() in combined.lower():
            return kw
    return "Thường"


def build_doc_list(state):
    docs = dict(state.get('documents', {}))
    for sid in state.get('seen_ids', []):
        sid_s = str(sid)
        if sid_s not in docs:
            docs[sid_s] = {
                'so_den': sid_s, 'so_ky_hieu': '', 'tac_gia': '',
                'trich_yeu': '', 'ngay_vb': '', 'ngay_den': '',
                'do_khan': 'Thường', 'but_phe': '', 'unit': '',
                'status': 'new', 'status_updated_at': None, 'note': '',
                'first_seen': None,
            }
    return docs


def do_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    state = load_state()
    docs = build_doc_list(state)

    os.makedirs(EXPORT_DIR, exist_ok=True)
    now = datetime.datetime.now()
    fname = f"cong-van-den_{now.strftime('%Y-W%V')}.xlsx"
    fpath = os.path.join(EXPORT_DIR, fname)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Công văn đến"

    headers = ["STT", "Số đến", "Ngày VB", "Ngày đến", "Số/Ký hiệu",
               "Tác giả", "Trích yếu", "Độ khẩn", "Đơn vị", "Trạng thái",
               "Cập nhật", "Ghi chú", "Bút phê", "File đính kèm"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    sorted_docs = sorted(docs.values(),
                         key=lambda d: int(d.get('so_den', '0') or '0'), reverse=True)

    for i, doc in enumerate(sorted_docs, 1):
        so_den = doc.get('so_den', '')
        attachments = state.get('documents', {}).get(so_den, {}).get('attachments', [])
        attach_str = ', '.join(a.get('display_name', '') for a in attachments) if attachments else ''
        row = [
            i,
            so_den,
            doc.get('ngay_vb', ''),
            doc.get('ngay_den', ''),
            doc.get('so_ky_hieu', ''),
            doc.get('tac_gia', ''),
            doc.get('trich_yeu', ''),
            get_urgency(doc),
            doc.get('unit', ''),
            doc.get('status', 'new'),
            doc.get('status_updated_at', ''),
            doc.get('note', ''),
            doc.get('but_phe', ''),
            attach_str,
        ]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=i + 1, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            if doc.get('status') == 'done':
                cell.font = Font(color="808080", italic=True)

    col_widths = [5, 8, 12, 12, 18, 25, 50, 10, 10, 10, 18, 25, 30, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = 'A2'

    wb.save(fpath)
    print(f"✅ Đã xuất Excel: {fpath} ({len(sorted_docs)} VB)")


def do_report(period):
    state = load_state()
    docs = build_doc_list(state)
    now = datetime.datetime.now()

    if period == 'weekly':
        start = now - datetime.timedelta(days=now.weekday() + 7)
        end = start + datetime.timedelta(days=7)
        label = start.strftime('%d/%m') + '-' + end.strftime('%d/%m/%Y')
    else:
        start = now.replace(day=1)
        if now.month == 12:
            end = now.replace(year=now.year + 1, month=1, day=1)
        else:
            end = now.replace(month=now.month + 1, day=1)
        label = now.strftime('%m/%Y')

    total = len(docs)
    status_count = {'new': 0, 'read': 0, 'wip': 0, 'done': 0, 'overdue': 0}
    urgent_count = 0
    new_in_period = 0
    unit_counts = {}

    for doc in docs.values():
        s = doc.get('status', 'new')
        status_count[s] = status_count.get(s, 0) + 1
        if get_urgency(doc) != "Thường":
            urgent_count += 1
        u = doc.get('unit', '').strip()
        if u:
            unit_counts[u] = unit_counts.get(u, 0) + 1
        fs = doc.get('first_seen')
        if fs:
            try:
                fs_dt = datetime.datetime.strptime(fs, '%Y-%m-%d %H:%M:%S')
                if start <= fs_dt < end:
                    new_in_period += 1
            except: pass

    done_pct = round(status_count['done'] / total * 100) if total else 0

    lines = [f"📊 Báo cáo {'tuần' if period == 'weekly' else 'tháng'} {label}"]
    lines.append(f"--{'--' * 30}")
    lines.append(f"📋 Tổng VB theo dõi: {total}")
    lines.append(f"🆕 Mới trong kỳ: {new_in_period}")
    lines.append(f"✅ Đã xử lý (done): {status_count['done']} ({done_pct}%)")
    if status_count['wip']:
        lines.append(f"⏳ Đang xử lý: {status_count['wip']}")
    if status_count['read']:
        lines.append(f"👁️ Đã đọc: {status_count['read']}")
    if status_count.get('overdue', 0):
        lines.append(f"❌ Quá hạn: {status_count['overdue']}")
    if status_count['new']:
        lines.append(f"📬 Chưa xử lý: {status_count['new']}")
    if urgent_count:
        lines.append(f"🔥 Văn bản khẩn: {urgent_count}")
    if unit_counts:
        unit_str = ", ".join(f"{u}={c}" for u, c in sorted(unit_counts.items()))
        lines.append(f"🏢 Theo đơn vị: {unit_str}")
    lines.append(f"🔗 {state.get('_source_url', 'https://congchuc.quangninh.gov.vn/Default.aspx?tabid=1126')}")
    lines.append("")
    lines.append("📌 VB chưa xử lý (new/read):")
    pending = sorted(
        [d for d in docs.values() if d.get('status', 'new') in ('new', 'read') and d.get('so_den', '')],
        key=lambda d: int(d.get('so_den', '0') or '0'), reverse=True
    )[:5]
    for d in pending:
        so_den = d.get('so_den', '?')
        skh = d.get('so_ky_hieu', '')
        ty = d.get('trich_yeu', '')[:60]
        u = d.get('unit', '')
        u_tag = f" [{u}]" if u else ""
        urg = "🔴" if get_urgency(d) != "Thường" else "  "
        lines.append(f"  {urg} #{so_den} {skh}: {ty}{u_tag}")

    print('\n'.join(lines))


def do_check_overdue():
    state = load_state()
    docs = state.get('documents', {})
    now = datetime.datetime.now()
    cutoff = now - datetime.timedelta(days=OVERDUE_DAYS)
    now_ts = now.strftime('%Y-%m-%d %H:%M:%S')
    newly_overdue = []

    for sid, doc in docs.items():
        status = doc.get('status', 'new')
        if status not in ('new', 'read', ''):
            continue
        ref_ts = doc.get('first_seen') or doc.get('status_updated_at')
        if not ref_ts:
            continue
        try:
            ref_dt = datetime.datetime.strptime(ref_ts, '%Y-%m-%d %H:%M:%S')
        except: continue
        if ref_dt >= cutoff:
            continue
        doc['status'] = 'overdue'
        doc['status_updated_at'] = now_ts
        existing_note = doc.get('note', '') or ''
        if '[Tự động đánh dấu quá hạn]' not in existing_note:
            doc['note'] = existing_note + '[Tự động đánh dấu quá hạn] "'
        newly_overdue.append((sid, doc.get('so_ky_hieu', ''), doc.get('trich_yeu', '')[:60]))

    if newly_overdue:
        state['last_check'] = now_ts
        save_state(state)
        print(f"❌ {len(newly_overdue)} VB quá hạn (> {OVERDUE_DAYS} ngày)\n")
        for sid, skh, ty in newly_overdue:
            print(f"  #{sid} {skh}: {ty}")
    else:
        print(f"✅ Không có VB quá hạn nào (> {OVERDUE_DAYS} ngày)")


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    mode = sys.argv[1]
    if mode == '--excel':
        do_excel()
    elif mode == '--weekly':
        do_report('weekly')
    elif mode == '--monthly':
        do_report('monthly')
    elif mode == '--check-overdue':
        do_check_overdue()
    elif mode in ('-h', '--help'):
        print(__doc__)
    else:
        print(f"Unknown mode: {mode}")
        print(__doc__)
        sys.exit(1)


if __name__ == '__main__':
    main()
