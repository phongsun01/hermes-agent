import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

def export_xlsx(merged_md_path):
    if not os.path.exists(merged_md_path):
        print(f"Lỗi: Không tìm thấy file {merged_md_path}")
        return

    output_dir = os.path.dirname(merged_md_path)
    # Lấy tên file gốc
    filename_without_ext = os.path.splitext(os.path.basename(merged_md_path))[0]
    merged_xlsx_path = os.path.join(output_dir, f"{filename_without_ext}.xlsx")

    all_data = []
    header = None

    with open(merged_md_path, 'r', encoding='utf-8') as f:
        lines = [line.strip() for line in f.readlines() if line.strip()]

    start_idx = -1
    for i, line in enumerate(lines):
        if line.startswith('|'):
            start_idx = i
            break
            
    if start_idx == -1:
        print("Lỗi: Không tìm thấy định dạng bảng (Markdown Table) trong file MD.")
        return

    hdr_line = lines[start_idx]
    header = [x.strip() for x in hdr_line.split('|')][1:-1]
    
    data_start_idx = start_idx + 2
    for line in lines[data_start_idx:]:
        if line.startswith('|'):
            cols = [x.strip() for x in line.split('|')][1:-1]
            if cols and len(cols) == len(header):
                all_data.append(cols)

    wb = Workbook()
    ws = wb.active
    ws.title = "Export Data"

    ws.append(header)
    
    for row in all_data:
        cleaned_row = []
        for i, val in enumerate(row):
            val_clean = val.replace(',', '').strip()
            # Xử lý cố gắng ép kiểu về Float/Int nếu nó là tiền tệ (Tự động nhận diện)
            if val_clean.replace('.','',1).isdigit() and " " not in val_clean:
                try:
                    if '.' in val_clean:
                        cleaned_row.append(float(val_clean))
                    else:
                        cleaned_row.append(int(val_clean))
                except:
                    cleaned_row.append(val)
            else:
                cleaned_row.append(val)
        ws.append(cleaned_row)

    # Styling for Excel (Premium Quality)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, size=12, color="FFFFFF")
    data_font = Font(size=11)
    
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Auto Adjust Widths logic (Simplified for aesthetics)
    for c in range(1, len(header) + 1):
        ws.column_dimensions[chr(64+c)].width = 25
    if len(header) >= 2:
        ws.column_dimensions['B'].width = 60 # Cột 2 thường chứa diễn giải nhiều nhất

    # Data formatting
    for r in range(2, len(all_data) + 2):
        for c in range(1, len(header) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = data_font
            cell.border = thin_border
            # Text align logic based on datatype
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '#,##0'
            else:
                cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    wb.save(merged_xlsx_path)
    print(f"[OPTIONAL-XLSX] Đã trích xuất và format Excel hoàn chỉnh: {merged_xlsx_path}")

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Sử dụng: python optional_export_xlsx.py <đường_dẫn_file_MERGED_MD>")
        sys.exit(1)
        
    md_file = sys.argv[1]
    export_xlsx(md_file)
